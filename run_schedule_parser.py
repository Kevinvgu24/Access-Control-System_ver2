import sys
import os
import re
import json
import shutil
import csv
from html.parser import HTMLParser

# Force stdout to UTF-8 to display Vietnamese characters correctly in Windows console
sys.stdout.reconfigure(encoding='utf-8')

class ExcelHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.raw_rows = []
        self.current_row = []
        self.in_cell = False
        self.cell_attrs = {}
        self.cell_text = []

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self.current_row = []
        elif tag in ("td", "th"):
            self.in_cell = True
            self.cell_attrs = dict(attrs)
            self.cell_text = []

    def handle_data(self, data):
        if self.in_cell:
            self.cell_text.append(data)

    def handle_endtag(self, tag):
        if tag in ("td", "th"):
            self.in_cell = False
            text = "".join(self.cell_text).strip()
            self.current_row.append({"text": text, "attrs": self.cell_attrs})
        elif tag == "tr":
            self.raw_rows.append(self.current_row)

class UniversalScheduleParser:
    def __init__(self, path, template_type='type1'):
        self.path = path.strip('"').strip("'")
        self.template_type = template_type
        self.grid = []
        self.timeline = []
        self.students = []
        self.parsed_records = []
        self.dominant_color = "NO_COLOR"
        self.is_xlsx = False

    def parse(self):
        # Determine file type
        if os.path.isdir(self.path):
            html_files = [f for f in os.listdir(self.path) if f.endswith('.html')]
            if html_files:
                target_file = os.path.join(self.path, html_files[0])
                self.is_xlsx = False
                if self.template_type == 'type1':
                    return self._parse_new_format(target_file)
                else:
                    return self._parse_html(target_file)
            else:
                raise ValueError("Không tìm thấy tệp .html trong thư mục.")
        elif self.path.endswith('.html') or self.path.endswith('.htm'):
            self.is_xlsx = False
            if self.template_type == 'type1':
                return self._parse_new_format(self.path)
            else:
                return self._parse_html(self.path)
        elif self.path.endswith('.xlsx'):
            self.is_xlsx = True
            if self.template_type == 'type1':
                return self._parse_new_format(self.path)
            else:
                return self._parse_xlsx()
        else:
            raise ValueError("Định dạng tệp không được hỗ trợ (chỉ nhận thư mục HTML, tệp .html hoặc tệp .xlsx).")

    def _parse_new_format(self, file_path_or_self_path):
        # 1. Build grid
        if self.is_xlsx:
            print(f"Đang đọc tệp Excel (.xlsx): {self.path}...")
            import uuid
            temp_file = f"temp_parsing_schedule_{uuid.uuid4().hex}.xlsx"
            shutil.copy(self.path, temp_file)
            
            try:
                import openpyxl
                wb = openpyxl.load_workbook(temp_file, data_only=True)
                sheet_name = None
                for name in wb.sheetnames:
                    if any(x in name.lower() for x in ['schedule', 'group', 'lịch', 'nhóm']):
                        sheet_name = name
                        break
                if sheet_name is None:
                    sheet_name = wb.sheetnames[0]
                    
                print(f"Đang phân tích trang tính (Sheet): '{sheet_name}'")
                ws = wb[sheet_name]
            except Exception as e:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                raise RuntimeError(f"Không thể mở tệp Excel: {e}")

            max_rows = ws.max_row
            max_cols = ws.max_column
            merged_ranges = ws.merged_cells.ranges
            
            # Read grid values and cell colors using the user's specific logic
            self.grid = [[{"text": "", "color": "NO_COLOR"} for _ in range(max_cols + 1)] for _ in range(max_rows + 1)]
            for r in range(1, max_rows + 1):
                for c in range(1, max_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    val = cell.value
                    txt = str(val).strip() if val is not None else ""
                    
                    if txt.endswith(".0"):
                        try:
                            txt = str(int(float(txt)))
                        except ValueError:
                            pass
                            
                    color = "NO_COLOR"
                    fill = cell.fill
                    is_colored = False
                    if fill and fill.fill_type == 'solid':
                        color_obj = fill.start_color
                        # Theme 1 thường là màu trắng mặc định của Excel
                        if hasattr(color_obj, 'theme') and color_obj.theme == 1:
                            is_colored = False
                        # Kiểm tra mã HEX (FFFFFFFF và 00000000 là trắng/trong suốt)
                        elif hasattr(color_obj, 'rgb') and color_obj.rgb not in ['FFFFFFFF', '00000000', '00FFFFFF', None]:
                            is_colored = True
                        # Nếu có theme khác 1 và khác None thì cũng tính là có màu
                        elif getattr(color_obj, 'theme', 1) not in [1, None]: 
                            is_colored = True

                    if is_colored:
                        color = "ACTIVE_COLOR"
                    self.grid[r][c] = {"text": txt, "color": color}

            # Apply merged values to all spanned cells
            for merged_range in merged_ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                origin_val = self.grid[min_row][min_col]
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        if r == min_row and c == min_col:
                            continue
                        self.grid[r][c]["text"] = origin_val["text"]
                        self.grid[r][c]["color"] = origin_val["color"]
            
            if os.path.exists(temp_file):
                os.remove(temp_file)
        else:
            print(f"Đang đọc tệp HTML: {file_path_or_self_path}...")
            with open(file_path_or_self_path, "r", encoding="utf-8") as f:
                html_content = f.read()

            parser = ExcelHTMLParser()
            parser.feed(html_content)
            raw_rows = parser.raw_rows

            max_cols = 0
            for row in raw_rows:
                cols_in_row = 0
                for cell in row:
                    cols_in_row += int(cell["attrs"].get("colspan", "1"))
                max_cols = max(max_cols, cols_in_row)

            self.grid = [[None for _ in range(max_cols)] for _ in range(len(raw_rows))]
            for r_idx, row in enumerate(raw_rows):
                c_idx = 0
                for cell in row:
                    while c_idx < max_cols and self.grid[r_idx][c_idx] is not None:
                        c_idx += 1
                    if c_idx >= max_cols:
                        break
                    colspan = int(cell["attrs"].get("colspan", "1"))
                    rowspan = int(cell["attrs"].get("rowspan", "1"))
                    for dr in range(rowspan):
                        for dc in range(colspan):
                            tr = r_idx + dr
                            tc = c_idx + dc
                            if tr < len(self.grid) and tc < max_cols:
                                self.grid[tr][tc] = cell
                    c_idx += colspan

        # 2. Extract year from filename/path
        year = 2026
        match = re.search(r'\b(202\d)\b', self.path)
        if match:
            year = int(match.group(1))

        # Helper to parse month names to numeric representation
        def parse_month_to_num(month_str):
            if not month_str:
                return 3
            month_str = str(month_str).lower().strip()
            digit_match = re.search(r'\d+', month_str)
            if digit_match:
                return int(digit_match.group(0))
            months_en = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
            for idx, m in enumerate(months_en):
                if m in month_str:
                    return idx + 1
            return 3

        # 3. Propagate and construct timeline
        self.timeline = []
        start_col = 5 if self.is_xlsx else 4
        
        # Row indices matching the new structure:
        # Row 7: Month | Row 8: Day | Row 9: Date | Row 10: M/A | Row 11: Session
        month_row = 7 if self.is_xlsx else 6
        d_row = 8 if self.is_xlsx else 7
        dt_row = 9 if self.is_xlsx else 8
        ma_row = 10 if self.is_xlsx else 9
        s_row = 11 if self.is_xlsx else 10
        
        months = []
        curr_month = "3"
        
        if len(self.grid) > month_row:
            row_len = len(self.grid[month_row])
            for c in range(start_col, row_len):
                cell = self.grid[month_row][c]
                txt = ""
                if cell:
                    txt = cell["text"].strip() if isinstance(cell, dict) else cell.get("text", "").strip()
                if txt:
                    curr_month = txt
                months.append(curr_month)

        for c in range(start_col, len(self.grid[0])):
            t_idx = c - start_col
            month_val = months[t_idx] if t_idx < len(months) else "3"
            
            day_of_week = ""
            if len(self.grid) > d_row and c < len(self.grid[d_row]):
                cell_d = self.grid[d_row][c]
                if cell_d:
                    day_of_week = cell_d["text"].strip() if isinstance(cell_d, dict) else cell_d.get("text", "").strip()

            day_val = ""
            if len(self.grid) > dt_row and c < len(self.grid[dt_row]):
                cell_dt = self.grid[dt_row][c]
                if cell_dt:
                    day_val = cell_dt["text"].strip() if isinstance(cell_dt, dict) else cell_dt.get("text", "").strip()
            if day_val.endswith('.0'):
                day_val = day_val[:-2]

            if not day_val or not day_val.isdigit():
                self.timeline.append(None)
                continue

            ma_val = ""
            if len(self.grid) > ma_row and c < len(self.grid[ma_row]):
                cell_ma = self.grid[ma_row][c]
                if cell_ma:
                    ma_val = cell_ma["text"].strip() if isinstance(cell_ma, dict) else cell_ma.get("text", "").strip()
            if ma_val.upper() == 'A':
                ma_val = 'Afternoon'
            elif ma_val.upper() == 'M':
                ma_val = 'Morning'

            session_val = ""
            if len(self.grid) > s_row and c < len(self.grid[s_row]):
                cell_s = self.grid[s_row][c]
                if cell_s:
                    session_val = cell_s["text"].strip() if isinstance(cell_s, dict) else cell_s.get("text", "").strip()
            if session_val.endswith('.0'):
                session_val = session_val[:-2]

            try:
                if '/' in day_val or '-' in day_val:
                    parts = re.split(r'[/|-]', day_val)
                    day_num = int(parts[0])
                    month_num = int(parts[1]) if len(parts) > 1 else parse_month_to_num(month_val)
                else:
                    day_num = int(day_val)
                    month_num = parse_month_to_num(month_val)
                date_str = f"{year:04d}-{month_num:02d}-{day_num:02d}"
            except Exception:
                date_str = f"{year:04d}-03-{int(day_val):02d}"

            self.timeline.append({
                "col": c,
                "date": date_str,
                "day": day_of_week,
                "ma": ma_val,
                "session": session_val
            })

        # 4. Parse students starting from row 13 (index 13 in xlsx, index 12 in html)
        start_row = 13 if self.is_xlsx else 12
        self.students = []
        
        curr_group = ""
        for r in range(start_row, len(self.grid)):
            row = self.grid[r]
            g_idx = 1 if self.is_xlsx else 0
            n_idx = 3 if self.is_xlsx else 2
            id_idx = 4 if self.is_xlsx else 3
            
            if len(row) <= max(g_idx, n_idx, id_idx):
                continue
                
            group_cell = row[g_idx]
            name_cell = row[n_idx]
            id_cell = row[id_idx]
            
            group_val = ""
            if group_cell:
                group_val = group_cell["text"].strip() if isinstance(group_cell, dict) else group_cell.get("text", "").strip()
            if group_val:
                curr_group = group_val

            name = ""
            if name_cell:
                name = name_cell["text"].strip() if isinstance(name_cell, dict) else name_cell.get("text", "").strip()
            std_id = ""
            if id_cell:
                std_id = id_cell["text"].strip() if isinstance(id_cell, dict) else id_cell.get("text", "").strip()
                
            if not name and not std_id:
                continue
                
            # Skip headers
            if name in ["Name ↓", "Họ và Tên", "MSSV", "STT"] or curr_group == "Group Nr.":
                continue
                
            if curr_group.endswith('.0'):
                curr_group = curr_group[:-2]
            if std_id.endswith('.0'):
                std_id = std_id[:-2]

            self.students.append({
                "row": r,
                "group": curr_group,
                "nr": "",
                "name": name,
                "id": std_id
            })

        # Helper for color matching
        def is_color_active(color_str):
            if not color_str:
                return False
            color_str = color_str.strip().lower()
            if color_str == "active_color":
                return True
            if color_str in ("no_color", "white", "#ffffff", "#fff", "ffffffff", "00000000", "00ffffff", "rgb(255,255,255)", "rgb(255, 255, 255)", "none", ""):
                return False
            return True

        # Parse active lab days for each student
        class_colors = {}
        if not self.is_xlsx:
            with open(file_path_or_self_path, "r", encoding="utf-8") as f:
                html_content = f.read()
            styles = re.findall(r'<style[^>]*>(.*?)</style>', html_content, re.DOTALL)
            if styles:
                style_content = styles[0]
                rules = re.findall(r'\.s(\d+)\b[^{]*\{([^}]+)\}', style_content)
                for cls_num, props in rules:
                    bg_match = re.search(r'background-color\s*:\s*([^;]+)', props)
                    if bg_match:
                        class_colors[f"s{cls_num}"] = bg_match.group(1).strip().lower()

        self.parsed_records = []
        for std in self.students:
            r = std["row"]
            group_nr = std["group"]
            name = std["name"]
            std_id = std["id"]
            row = self.grid[r]
            
            for c in range(start_col, len(row)):
                t_idx = c - start_col
                if t_idx >= len(self.timeline) or self.timeline[t_idx] is None:
                    continue
                    
                t_info = self.timeline[t_idx]
                cell = row[c]
                if not cell:
                    continue
                    
                cell_text = cell["text"].strip() if isinstance(cell, dict) else cell.get("text", "").strip()
                
                # Check color
                color = "NO_COLOR"
                if self.is_xlsx:
                    color = cell.get("color", "NO_COLOR")
                else:
                    style_val = cell["attrs"].get("style", "").lower() if isinstance(cell, dict) and "attrs" in cell else ""
                    bg_match = re.search(r'background(?:-color)?\s*:\s*([^;]+)', style_val)
                    if bg_match:
                        color = bg_match.group(1).strip().lower()
                    else:
                        cls = cell["attrs"].get("class", "") if isinstance(cell, dict) and "attrs" in cell else ""
                        color = class_colors.get(cls, "NO_COLOR")

                if is_color_active(color):
                    exp = cell_text
                    if not exp:
                        exp = "Có lịch (Ô gộp)"
                        
                    self.parsed_records.append({
                        "student_id": std_id,
                        "student_name": name,
                        "group_nr": group_nr,
                        "student_nr": "",
                        "date": t_info["date"],
                        "day_of_week": t_info["day"],
                        "ma": t_info["ma"],
                        "session_num": t_info["session"],
                        "experiment": exp
                    })
        
        # Write debug file
        try:
            with open("debug_parse.log", "w", encoding="utf-8") as df:
                df.write(f"Is XLSX: {self.is_xlsx}\n")
                df.write(f"Grid dimensions: {len(self.grid)} rows, {len(self.grid[0]) if self.grid else 0} cols\n")
                
                df.write("\n--- FIRST 20 ROWS & 10 COLS ---\n")
                for r in range(min(20, len(self.grid))):
                    row_cells = []
                    for c in range(min(15, len(self.grid[r]))):
                        cell = self.grid[r][c]
                        if cell:
                            txt = cell.get("text", "") if isinstance(cell, dict) else ""
                            col_val = cell.get("color", "") if isinstance(cell, dict) else ""
                            row_cells.append(f"[{r},{c}]={txt}({col_val})")
                        else:
                            row_cells.append(f"[{r},{c}]=None")
                    df.write(" | ".join(row_cells) + "\n")
                
                df.write(f"\nTimeline length: {len(self.timeline)}\n")
                for idx, t in enumerate(self.timeline):
                    df.write(f"Col {idx}: {t}\n")
                    
                df.write(f"\nStudents parsed: {len(self.students)}\n")
                for std in self.students[:50]:
                    df.write(f"Row {std['row']}: {std}\n")
                    
                df.write(f"\nParsed records: {len(self.parsed_records)}\n")
                for rec in self.parsed_records[:50]:
                    df.write(f"{rec}\n")
        except Exception as log_err:
            print(f"DEBUG LOG ERROR: {log_err}")

        return self.parsed_records

    def _parse_html(self, file_path):
        print(f"Đang đọc tệp HTML: {file_path}...")
        with open(file_path, "r", encoding="utf-8") as f:
            html_content = f.read()

        # Parse CSS stylesheet colors
        class_colors = {}
        styles = re.findall(r'<style[^>]*>(.*?)</style>', html_content, re.DOTALL)
        if styles:
            style_content = styles[0]
            rules = re.findall(r'\.s(\d+)\b[^{]*\{([^}]+)\}', style_content)
            for cls_num, props in rules:
                bg_match = re.search(r'background-color\s*:\s*([^;]+)', props)
                if bg_match:
                    bg_color = bg_match.group(1).strip().lower()
                    class_colors[f"s{cls_num}"] = bg_color

        parser = ExcelHTMLParser()
        parser.feed(html_content)
        raw_rows = parser.raw_rows

        # Build 2D grid
        max_cols = 0
        for row in raw_rows:
            cols_in_row = 0
            for cell in row:
                cols_in_row += int(cell["attrs"].get("colspan", "1"))
            max_cols = max(max_cols, cols_in_row)

        self.grid = [[None for _ in range(max_cols)] for _ in range(len(raw_rows))]
        for r_idx, row in enumerate(raw_rows):
            c_idx = 0
            for cell in row:
                while c_idx < max_cols and self.grid[r_idx][c_idx] is not None:
                    c_idx += 1
                if c_idx >= max_cols:
                    break
                colspan = int(cell["attrs"].get("colspan", "1"))
                rowspan = int(cell["attrs"].get("rowspan", "1"))
                for dr in range(rowspan):
                    for dc in range(colspan):
                        tr = r_idx + dr
                        tc = c_idx + dc
                        if tr < len(self.grid) and tc < max_cols:
                            self.grid[tr][tc] = cell
                c_idx += colspan

        # Detect row indices dynamically
        y_r, m_r, d_r, dt_r, ma_r, s_r, h_r = self._detect_row_indices(self.grid, len(self.grid), max_cols, False)
        if None in (y_r, m_r, d_r, dt_r, ma_r, s_r, h_r):
            raise ValueError("Không thể phát hiện đầy đủ cấu trúc các dòng lịch trình (Năm, Tháng, Ngày, Ca, Phiên, Header).")

        # Parse student rows (will populate self.group_col_idx, self.nr_col_idx, etc.)
        self._parse_students_rows(h_r, name_col=3, id_col=4, group_col=1, nr_col=2)
        
        # Dynamically detect where the timeline starts by looking for the first digit in the date row
        start_col = None
        for c in range(len(self.grid[dt_r])):
            cell = self.grid[dt_r][c]
            if not cell:
                continue
            txt = cell["text"].strip()
            if txt.endswith('.0'):
                txt = txt[:-2]
            if txt.isdigit():
                start_col = c
                break
        if start_col is None:
            start_col = 6 # fallback

        print(f"Phát hiện cột bắt đầu Timeline (HTML): {start_col}")

        # Parse timeline columns
        self._parse_timeline_columns(y_r, m_r, d_r, dt_r, ma_r, s_r, start_col=start_col)

        # Detect dominant color (blocked cell color) in student schedule grid
        color_freq = {}
        for std in self.students:
            r_idx = std["row"]
            row = self.grid[r_idx]
            for c_idx in range(start_col, min(142, len(row))):
                cell = row[c_idx]
                if not cell:
                    continue
                cls = cell["attrs"].get("class", "")
                color = class_colors.get(cls, "NO_COLOR")
                color_freq[color] = color_freq.get(color, 0) + 1
        self.dominant_color = max(color_freq, key=color_freq.get) if color_freq else "NO_COLOR"

        # Resolve experiments and build schedule records
        g_col = getattr(self, "group_col_idx", 1)
        group_experiments = {}
        for r_idx in range(h_r + 1, len(self.grid)):
            row = self.grid[r_idx]
            if len(row) <= g_col or not row[g_col]:
                continue
            group_nr = row[g_col]["text"]
            for c_idx in range(start_col, min(142, len(row))):
                cell = row[c_idx]
                if cell and cell["text"]:
                    group_experiments[(group_nr, c_idx)] = cell["text"]

        self.parsed_records = []
        for std in self.students:
            r_idx = std["row"]
            group_nr = std["group"]
            row = self.grid[r_idx]

            for c_idx in range(start_col, min(142, len(row))):
                t_idx = c_idx - start_col
                if t_idx >= len(self.timeline) or self.timeline[t_idx] is None:
                    continue

                t_info = self.timeline[t_idx]
                cell = row[c_idx]
                if not cell:
                    continue

                cls = cell["attrs"].get("class", "")
                color = class_colors.get(cls, "NO_COLOR")
                cell_text = cell["text"].strip()
                
                # Active if the cell has text OR color is different from the dominant blocked color
                is_active = False
                if cell_text:
                    is_active = True
                elif color != self.dominant_color:
                    if self.dominant_color == "NO_COLOR":
                        if color != "NO_COLOR":
                            is_active = True
                    else:
                        is_active = True
                
                if is_active:
                    exp = cell_text
                    if not exp:
                        exp = group_experiments.get((group_nr, c_idx), "")

                    self.parsed_records.append({
                        "student_id": std["id"],
                        "student_name": std["name"],
                        "group_nr": group_nr,
                        "student_nr": std["nr"],
                        "date": t_info["date"],
                        "day_of_week": t_info["day"],
                        "ma": t_info["ma"],
                        "session_num": t_info["session"],
                        "experiment": exp
                    })
        return self.parsed_records

    def _parse_xlsx(self):
        print(f"Đang đọc tệp Excel (.xlsx): {self.path}...")
        import uuid
        temp_file = f"temp_parsing_schedule_{uuid.uuid4().hex}.xlsx"
        shutil.copy(self.path, temp_file)
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(temp_file, data_only=True)
            
            # Select correct sheet dynamically by scanning sheet names
            sheet_name = None
            for name in wb.sheetnames:
                if any(x in name.lower() for x in ['schedule', 'group', 'lịch', 'nhóm']):
                    sheet_name = name
                    break
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]
                
            print(f"Đang phân tích trang tính (Sheet): '{sheet_name}'")
            ws = wb[sheet_name]
        except Exception as e:
            if os.path.exists(temp_file):
                os.remove(temp_file)
            raise RuntimeError(f"Không thể mở tệp Excel: {e}. Vui lòng cài đặt openpyxl bằng: pip install openpyxl")

        max_rows = ws.max_row
        max_cols = ws.max_column
        merged_ranges = ws.merged_cells.ranges
        
        # Read grid values and cell colors
        self.grid = [[{"text": "", "color": "NO_COLOR"} for _ in range(max_cols + 1)] for _ in range(max_rows + 1)]
        for r in range(1, max_rows + 1):
            for c in range(1, max_cols + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                txt = str(val).strip() if val is not None else ""
                
                # Clean floats generated by Excel (like 102240004.0 to 102240004)
                if txt.endswith(".0"):
                    try:
                        txt = str(int(float(txt)))
                    except ValueError:
                        pass
                        
                color = "NO_COLOR"
                fill = cell.fill
                if fill and fill.fill_type and fill.fill_type != 'none':
                    if hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
                        color = str(fill.fgColor.rgb)
                self.grid[r][c] = {"text": txt, "color": color}

        # Apply merged values to all spanned cells
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            origin_val = self.grid[min_row][min_col]
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if r == min_row and c == min_col:
                        continue
                    self.grid[r][c]["text"] = origin_val["text"]

        # Detect row indices dynamically
        y_r, m_r, d_r, dt_r, ma_r, s_r, h_r = self._detect_row_indices(self.grid, max_rows, max_cols, True)
        if None in (y_r, m_r, d_r, dt_r, ma_r, s_r, h_r):
            if os.path.exists(temp_file): os.remove(temp_file)
            raise ValueError("Không thể phát hiện cấu trúc các dòng lịch trình (Năm, Tháng, Ngày, Ca, Phiên, Header).")

        # Parse student rows (will populate self.group_col_idx, etc.)
        self._parse_students_rows(h_r, name_col=3, id_col=4, group_col=1, nr_col=2)

        # Dynamically detect where the timeline starts by looking for the first digit in the date row
        start_col = None
        for c in range(1, len(self.grid[dt_r])):
            cell = self.grid[dt_r][c]
            if not cell:
                continue
            txt = cell["text"].strip()
            if txt.endswith('.0'):
                txt = txt[:-2]
            if txt.isdigit():
                start_col = c
                break
        if start_col is None:
            start_col = 7 # fallback

        print(f"Phát hiện cột bắt đầu Timeline (XLSX): {start_col}")

        # Parse timeline columns starting from start_col
        self._parse_timeline_columns(y_r, m_r, d_r, dt_r, ma_r, s_r, start_col=start_col)

        # Detect dominant color in grid
        color_freq = {}
        for std in self.students:
            r_idx = std["row"]
            for c_idx in range(start_col, min(143, max_cols + 1)):
                cell = self.grid[r_idx][c_idx]
                color = cell["color"]
                color_freq[color] = color_freq.get(color, 0) + 1
        self.dominant_color = max(color_freq, key=color_freq.get) if color_freq else "NO_COLOR"

        # Resolve experiments and build schedule records
        g_col = getattr(self, "group_col_idx", 1)
        group_experiments = {}
        for r_idx in range(h_r + 1, max_rows + 1):
            if len(self.grid[r_idx]) <= g_col:
                continue
            group_nr = self.grid[r_idx][g_col]["text"]
            if not group_nr or group_nr in ["MSE", "ICT", "Name ↓"]:
                continue
            for c_idx in range(start_col, min(143, max_cols + 1)):
                cell = self.grid[r_idx][c_idx]
                if cell and cell["text"]:
                    group_experiments[(group_nr, c_idx)] = cell["text"]

        self.parsed_records = []
        for std in self.students:
            r_idx = std["row"]
            group_nr = std["group"]

            for c_idx in range(start_col, min(143, max_cols + 1)):
                t_idx = c_idx - start_col
                if t_idx >= len(self.timeline) or self.timeline[t_idx] is None:
                    continue

                t_info = self.timeline[t_idx]
                cell = self.grid[r_idx][c_idx]
                cell_text = cell["text"].strip()
                
                # Active if the cell has text OR color is different from the dominant blocked color
                is_active = False
                if cell_text:
                    is_active = True
                elif cell["color"] != self.dominant_color:
                    if self.dominant_color == "NO_COLOR":
                        if cell["color"] != "NO_COLOR":
                            is_active = True
                    else:
                        is_active = True

                if is_active:
                    exp = cell_text
                    if not exp:
                        exp = group_experiments.get((group_nr, c_idx), "")

                    self.parsed_records.append({
                        "student_id": std["id"],
                        "student_name": std["name"],
                        "group_nr": group_nr,
                        "student_nr": std["nr"],
                        "date": t_info["date"],
                        "day_of_week": t_info["day"],
                        "ma": t_info["ma"],
                        "session_num": t_info["session"],
                        "experiment": exp
                    })

        if os.path.exists(temp_file):
            os.remove(temp_file)
            
        return self.parsed_records

    def _detect_row_indices(self, grid, max_rows, max_cols, is_xlsx):
        """Find Year, Month, Day, Date, Session, and Header row indices dynamically."""
        y_r = m_r = d_r = dt_r = ma_r = s_r = h_r = None
        limit_col = min(7, max_cols)
        
        start_idx = 1 if is_xlsx else 0
        end_idx = max_rows + 1 if is_xlsx else max_rows
        
        for r in range(start_idx, end_idx):
            if is_xlsx:
                row_txts = [grid[r][c]["text"].strip().lower() for c in range(1, limit_col + 1)]
            else:
                row = grid[r]
                if not row: continue
                row_txts = [cell["text"].strip().lower() if cell else "" for cell in row[:limit_col]]
                
            if any("year" in t or "năm" in t or "nam" in t for t in row_txts) and y_r is None: y_r = r
            if any("month" in t or "tháng" in t or "thang" in t for t in row_txts) and m_r is None: m_r = r
            if any("day" in t or "thứ" in t or "thu" in t for t in row_txts) and d_r is None: d_r = r
            if any("date" in t or "ngày" in t or "ngay" in t for t in row_txts) and dt_r is None: dt_r = r
            if any("morning/afternoon" in t or "(m/a)" in t or "buổi" in t or "buoi" in t for t in row_txts) and ma_r is None: ma_r = r
            if any("session" in t or "ca" in t or "phiên" in t or "phien" in t for t in row_txts) and s_r is None: s_r = r
            if any("group nr" in t or "name ↓" in t or "id ↓" in t or "nhóm" in t or "nhom" in t or "họ và tên" in t or "ho va ten" in t or "mssv" in t or "stt" in t for t in row_txts) and h_r is None: h_r = r
            
        return y_r, m_r, d_r, dt_r, ma_r, s_r, h_r

    def _parse_timeline_columns(self, y_r, m_r, d_r, dt_r, ma_r, s_r, start_col):
        """Construct date and ca information for each timeline column."""
        def get_texts(r):
            if self.is_xlsx:
                return [self.grid[r][c]["text"] for c in range(1, len(self.grid[r]))]
            else:
                return [self.grid[r][c]["text"] if self.grid[r][c] else "" for c in range(len(self.grid[r]))]
                
        years_raw = get_texts(y_r)
        months_raw = get_texts(m_r)
        days_raw = get_texts(d_r)
        dates_raw = get_texts(dt_r)
        mas_raw = get_texts(ma_r)
        sessions_raw = get_texts(s_r)
        
        def propagate(lst):
            res = []
            curr = ""
            for val in lst:
                if val: curr = val
                res.append(curr)
            return res

        # Year and Month are merged headers, so propagate them
        years = propagate(years_raw)
        months = propagate(months_raw)
        # Day, Date, M/A and Session are column-specific, do NOT propagate
        days = days_raw
        dates = dates_raw
        mas = mas_raw
        sessions = sessions_raw

        offset = 1 if self.is_xlsx else 0
        
        self.timeline = []
        for c in range(start_col, len(self.grid[0]) if not self.is_xlsx else len(self.grid[0])):
            idx = c - offset
            if idx >= len(dates):
                self.timeline.append(None)
                continue
            date_val = dates[idx]
            
            # Clean floating representations in Excel like '23.0' to '23'
            if date_val.endswith('.0'):
                try: date_val = str(int(float(date_val)))
                except ValueError: pass
                
            if not date_val or not date_val.isdigit():
                self.timeline.append(None)
                continue
                
            day_num = int(date_val)
            year_val = years[idx]
            if year_val.endswith('.0'):
                try: year_val = str(int(float(year_val)))
                except ValueError: pass
            month_val = months[idx]
            if month_val.endswith('.0'):
                try: month_val = str(int(float(month_val)))
                except ValueError: pass
                
            year = int(year_val) if year_val.isdigit() else 2026
            if year < 100:
                year += 2000
                
            month = int(month_val) if month_val.isdigit() else 3
            date_str = f"{year:04d}-{month:02d}-{day_num:02d}"

            s_val = sessions[idx]
            if s_val.endswith('.0'):
                try: s_val = str(int(float(s_val)))
                except ValueError: pass

            self.timeline.append({
                "col": c,
                "date": date_str,
                "day": days[idx],
                "ma": mas[idx],
                "session": s_val
            })

    def _parse_students_rows(self, h_r, name_col, id_col, group_col, nr_col):
        """Parse student info (skipping sub-headers)."""
        header_row = self.grid[h_r]
        start_col_idx = 1 if self.is_xlsx else 0
        
        detected_group = detected_nr = detected_name = detected_id = None
        for c_idx in range(start_col_idx, len(header_row)):
            cell = header_row[c_idx]
            if not cell:
                continue
            text = cell.get("text", "").strip().lower() if isinstance(cell, dict) else str(cell).strip().lower()
            if not text:
                continue
            
            if any(x in text for x in ["group", "nhóm", "nhom"]) and detected_group is None:
                detected_group = c_idx
            elif any(x in text for x in ["nr", "no.", "stt", "number", "thứ tự", "thu tu"]) and detected_nr is None:
                detected_nr = c_idx
            elif any(x in text for x in ["name", "tên", "ten", "họ", "ho"]) and detected_name is None:
                detected_name = c_idx
            elif any(x in text for x in ["id", "mssv", "mã", "ma"]) and detected_id is None:
                detected_id = c_idx
                
        # Assign to self so other methods can access the correct indices
        self.group_col_idx = detected_group if detected_group is not None else group_col
        self.nr_col_idx = detected_nr if detected_nr is not None else nr_col
        self.name_col_idx = detected_name if detected_name is not None else name_col
        self.id_col_idx = detected_id if detected_id is not None else id_col
        
        print(f"Phát hiện cột: Group={self.group_col_idx}, Nr={self.nr_col_idx}, Name={self.name_col_idx}, ID={self.id_col_idx}")
        
        self.students = []
        start_idx = h_r + 1
        end_idx = len(self.grid)
        
        for r in range(start_idx, end_idx):
            row = self.grid[r]
            if not row or len(row) <= max(self.group_col_idx, self.nr_col_idx, self.name_col_idx, self.id_col_idx):
                continue
            
            group_nr = row[self.group_col_idx]["text"] if row[self.group_col_idx] else ""
            student_nr = row[self.nr_col_idx]["text"] if row[self.nr_col_idx] else ""
            name = row[self.name_col_idx]["text"] if row[self.name_col_idx] else ""
            std_id = row[self.id_col_idx]["text"] if row[self.id_col_idx] else ""

            if not name and not std_id: continue
            if name in ["Name ↓", "MSE", "ICT", "Nhóm", "STT", "Họ và Tên", "MSSV"] or name.startswith("Students:"): continue
            
            # Clean floating representation of student numbers and IDs
            if group_nr.endswith('.0'):
                try: group_nr = str(int(float(group_nr)))
                except ValueError: pass
            if student_nr.endswith('.0'):
                try: student_nr = str(int(float(student_nr)))
                except ValueError: pass
            if std_id.endswith('.0'):
                try: std_id = str(int(float(std_id)))
                except ValueError: pass
                
            self.students.append({
                "row": r,
                "group": group_nr,
                "nr": student_nr,
                "name": name,
                "id": std_id
            })

    def _build_grid(self, file_path_or_self_path):
        import re
        if self.is_xlsx:
            import shutil
            import uuid
            import os
            temp_file = f"temp_parsing_schedule_{uuid.uuid4().hex}.xlsx"
            shutil.copy(file_path_or_self_path, temp_file)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(temp_file, data_only=True)
                sheet_name = None
                for name in wb.sheetnames:
                    if any(x in name.lower() for x in ['schedule', 'group', 'lịch', 'nhóm']):
                        sheet_name = name
                        break
                if sheet_name is None:
                    sheet_name = wb.sheetnames[0]
                ws = wb[sheet_name]
                max_rows = ws.max_row
                max_cols = ws.max_column
                merged_ranges = ws.merged_cells.ranges
                
                self.grid = [[{"text": "", "color": "NO_COLOR"} for _ in range(max_cols)] for _ in range(max_rows)]
                for r in range(1, max_rows + 1):
                    for c in range(1, max_cols + 1):
                        cell = ws.cell(row=r, column=c)
                        val = cell.value
                        txt = str(val).strip() if val is not None else ""
                        if txt.endswith(".0"):
                            try: txt = str(int(float(txt)))
                            except ValueError: pass
                        color = "NO_COLOR"
                        fill = cell.fill
                        if fill and fill.fill_type and fill.fill_type != 'none':
                            if hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
                                color = str(fill.fgColor.rgb)
                        self.grid[r-1][c-1] = {"text": txt, "color": color}
                        
                for merged_range in merged_ranges:
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    # Ensure indices are within bounds
                    m_row = min(min_row - 1, len(self.grid) - 1)
                    m_col = min(min_col - 1, len(self.grid[0]) - 1)
                    origin_val = self.grid[m_row][m_col]
                    for r in range(min_row, min(max_row + 1, len(self.grid) + 1)):
                        for c in range(min_col, min(max_col + 1, len(self.grid[0]) + 1)):
                            if r == min_row and c == min_col:
                                continue
                            self.grid[r-1][c-1]["text"] = origin_val["text"]
                            self.grid[r-1][c-1]["color"] = origin_val["color"]
            finally:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
        else:
            with open(file_path_or_self_path, "r", encoding="utf-8") as f:
                html_content = f.read()
            parser = ExcelHTMLParser()
            parser.feed(html_content)
            raw_rows = parser.raw_rows
            max_cols = 0
            for row in raw_rows:
                cols_in_row = 0
                for cell in row:
                    cols_in_row += int(cell["attrs"].get("colspan", "1"))
                max_cols = max(max_cols, cols_in_row)
            self.grid = [[None for _ in range(max_cols)] for _ in range(len(raw_rows))]
            for r_idx, row in enumerate(raw_rows):
                c_idx = 0
                for cell in row:
                    while c_idx < max_cols and self.grid[r_idx][c_idx] is not None:
                        c_idx += 1
                    if c_idx >= max_cols:
                        break
                    colspan = int(cell["attrs"].get("colspan", "1"))
                    rowspan = int(cell["attrs"].get("rowspan", "1"))
                    for dr in range(rowspan):
                        for dc in range(colspan):
                            tr = r_idx + dr
                            tc = c_idx + dc
                            if tr < len(self.grid) and tc < max_cols:
                                self.grid[tr][tc] = cell
                    c_idx += colspan
            
            # Convert HTML cells to standard text/color dicts
            class_colors = {}
            styles = re.findall(r'<style[^>]*>(.*?)</style>', html_content, re.DOTALL)
            if styles:
                style_content = styles[0]
                rules = re.findall(r'\.s(\d+)\b[^{]*\{([^}]+)\}', style_content)
                for cls_num, props in rules:
                    bg_match = re.search(r'background-color\s*:\s*([^;]+)', props)
                    if bg_match:
                        class_colors[f"s{cls_num}"] = bg_match.group(1).strip().lower()

            for r in range(len(self.grid)):
                for c in range(len(self.grid[r])):
                    cell = self.grid[r][c]
                    if cell is None:
                        self.grid[r][c] = {"text": "", "color": "NO_COLOR"}
                    else:
                        txt = cell.get("text", "").strip()
                        style_val = cell["attrs"].get("style", "").lower() if "attrs" in cell else ""
                        bg_match = re.search(r'background(?:-color)?\s*:\s*([^;]+)', style_val)
                        color = "NO_COLOR"
                        if bg_match:
                            color = bg_match.group(1).strip().lower()
                        else:
                            cls = cell["attrs"].get("class", "") if "attrs" in cell else ""
                            color = class_colors.get(cls, "NO_COLOR")
                        self.grid[r][c] = {"text": txt, "color": color}

    def parse_with_mapping(self, mapping):
        import re
        self._build_grid(self.path)
        
        year = 2026
        match = re.search(r'\b(202\d)\b', self.path)
        if match:
            year = int(match.group(1))

        month_row = int(mapping.get("month_row", 5))
        day_of_week_row = int(mapping.get("day_of_week_row", 6))
        date_row = int(mapping.get("date_row", 7))
        ma_row = int(mapping.get("ma_row", 8))
        session_row = int(mapping.get("session_row", 9))
        
        group_col = int(mapping.get("group_col", 0))
        name_col = int(mapping.get("name_col", 2))
        id_col = int(mapping.get("id_col", 3))
        
        start_col = int(mapping.get("start_col", 4))
        start_row = int(mapping.get("start_row", 12))

        self.timeline = []
        months = []
        curr_month = "4"
        
        if len(self.grid) > month_row:
            row_len = len(self.grid[month_row])
            for c in range(start_col, row_len):
                cell = self.grid[month_row][c]
                txt = ""
                if cell:
                    txt = cell.get("text", "").strip()
                if txt:
                    m_match = re.search(r'\d+', txt)
                    curr_month = m_match.group(0) if m_match else txt
                months.append(curr_month)

        for c in range(start_col, len(self.grid[0])):
            t_idx = c - start_col
            month_val = months[t_idx] if t_idx < len(months) else "4"
            
            day_of_week = ""
            if len(self.grid) > day_of_week_row and c < len(self.grid[day_of_week_row]):
                cell_d = self.grid[day_of_week_row][c]
                if cell_d:
                    day_of_week = cell_d.get("text", "").strip()

            day_val = ""
            if len(self.grid) > date_row and c < len(self.grid[date_row]):
                cell_dt = self.grid[date_row][c]
                if cell_dt:
                    day_val = cell_dt.get("text", "").strip()
            if day_val.endswith('.0'):
                day_val = day_val[:-2]

            if not day_val or not day_val.isdigit():
                self.timeline.append(None)
                continue

            ma_val = ""
            if len(self.grid) > ma_row and c < len(self.grid[ma_row]):
                cell_ma = self.grid[ma_row][c]
                if cell_ma:
                    ma_val = cell_ma.get("text", "").strip()
            if ma_val.upper() == 'A':
                ma_val = 'Afternoon'
            elif ma_val.upper() == 'M':
                ma_val = 'Morning'

            session_val = ""
            if len(self.grid) > session_row and c < len(self.grid[session_row]):
                cell_s = self.grid[session_row][c]
                if cell_s:
                    session_val = cell_s.get("text", "").strip()
            if session_val.endswith('.0'):
                session_val = session_val[:-2]

            try:
                month_num = int(month_val) if month_val.isdigit() else 4
                day_num = int(day_val)
                date_str = f"{year:04d}-{month_num:02d}-{day_num:02d}"
            except Exception:
                date_str = f"2026-04-{int(day_val):02d}"

            self.timeline.append({
                "col": c,
                "date": date_str,
                "day": day_of_week,
                "ma": ma_val,
                "session": session_val
            })

        self.students = []
        for r in range(start_row, len(self.grid)):
            row = self.grid[r]
            if len(row) <= max(group_col, name_col, id_col):
                continue
                
            group_cell = row[group_col]
            name_cell = row[name_col]
            id_cell = row[id_col]
            
            group_nr = group_cell.get("text", "").strip() if group_cell else ""
            name = name_cell.get("text", "").strip() if name_cell else ""
            std_id = id_cell.get("text", "").strip() if id_cell else ""
                
            if not name and not std_id:
                continue
                
            if name in ["Name ↓", "Họ và Tên", "MSSV", "STT"] or group_nr == "Group Nr.":
                continue
                
            if group_nr.endswith('.0'):
                group_nr = group_nr[:-2]
            if std_id.endswith('.0'):
                std_id = std_id[:-2]

            self.students.append({
                "row": r,
                "group": group_nr,
                "nr": "",
                "name": name,
                "id": std_id
            })

        def is_color_active(color_str):
            if not color_str:
                return False
            color_str = color_str.strip().lower()
            if color_str in ("no_color", "white", "#ffffff", "#fff", "ffffffff", "00000000", "00ffffff", "rgb(255,255,255)", "rgb(255, 255, 255)", "none", ""):
                return False
            return True

        self.parsed_records = []
        for std in self.students:
            r = std["row"]
            group_nr = std["group"]
            name = std["name"]
            std_id = std["id"]
            row = self.grid[r]
            
            for c in range(start_col, len(row)):
                t_idx = c - start_col
                if t_idx >= len(self.timeline) or self.timeline[t_idx] is None:
                    continue
                    
                t_info = self.timeline[t_idx]
                cell = row[c]
                if not cell:
                    continue
                    
                cell_text = cell.get("text", "").strip()
                color = cell.get("color", "NO_COLOR")

                if is_color_active(color):
                    exp = cell_text
                    if not exp:
                        exp = "Lab Session"
                        
                    self.parsed_records.append({
                        "student_id": std_id,
                        "student_name": name,
                        "group_nr": group_nr,
                        "student_nr": "",
                        "date": t_info["date"],
                        "day_of_week": t_info["day"],
                        "ma": t_info["ma"],
                        "session_num": t_info["session"],
                        "experiment": exp
                    })
        return self.parsed_records

def print_menu():
    print("\n" + "="*50)
    print("      LAB SCHEDULE LOOKUP SYSTEM      ")
    print("="*50)
    print(" [1] Display all lab schedules (First 30 rows)")
    print(" [2] Search schedule by Student Name or ID")
    print(" [3] Search lab schedule by Date (YYYY-MM-DD)")
    print(" [4] Export all data to CSV file")
    print(" [5] Exit program")
    print("="*50)

def main():
    default_path = r"c:\Users\Phant\Downloads\ECE2024_ICT & MSE major_Com. Eng Lab_Schedule & Group_20260401.xlsx"
    
    print("Welcome to Lab Schedule Parser Tool.")
    print(f"Default path suggestion:\n  {default_path}")
    path_input = input("Enter file path (or press Enter for default): ").strip()
    
    if not path_input:
        path_input = default_path

    if not os.path.exists(path_input):
        print(f"Error: Path '{path_input}' does not exist. Please try again.")
        return

    try:
        parser = UniversalScheduleParser(path_input)
        records = parser.parse()
        print(f"\nParsing successful! Found:")
        print(f"  - Total students: {len(parser.students)}")
        print(f"  - Total lab visits: {len(records)}")
    except Exception as e:
        print(f"\nError parsing file: {e}")
        return

    while True:
        print_menu()
        choice = input("Your choice (1-5): ").strip()
        
        if choice == '1':
            print(f"\n--- FIRST 30 LAB VISITS (Total: {len(records)}) ---")
            print(f"{'ID':<10} | {'Full Name':<25} | {'Group':<5} | {'Date':<10} | {'Day':<4} | {'M/A':<5} | {'Sess':<5} | {'Experiment'}")
            print("-" * 90)
            for r in records[:30]:
                print(f"{r['student_id']:<10} | {r['student_name']:<25} | {r['group_nr']:<5} | {r['date']:<10} | {r['day_of_week']:<4} | {r['ma']:<5} | {r['session_num']:<5} | {r['experiment']}")
            if len(records) > 30:
                print(f"... and {len(records) - 30} more entries.")

        elif choice == '2':
            query = input("Enter Student Name or ID to search: ").strip().lower()
            if not query:
                continue
            
            results = [r for r in records if query in r['student_name'].lower() or query == r['student_id'].lower()]
            
            if not results:
                print(f"\nNo schedule found for student: '{query}'")
            else:
                print(f"\n--- STUDENT LOOKUP RESULT: {results[0]['student_name']} (ID: {results[0]['student_id']}) ---")
                print(f"{'Date':<10} | {'Day':<4} | {'M/A':<5} | {'Sess':<5} | {'Experiment'}")
                print("-" * 50)
                for r in results:
                    print(f"{r['date']:<10} | {r['day_of_week']:<4} | {r['ma']:<5} | {r['session_num']:<5} | {r['experiment']}")

        elif choice == '3':
            date_query = input("Enter date to search (Format YYYY-MM-DD, e.g. 2026-04-17): ").strip()
            if not date_query:
                continue
            
            results = [r for r in records if r['date'] == date_query]
            
            if not results:
                print(f"\nNo students scheduled for lab on '{date_query}'.")
            else:
                print(f"\n--- LAB ATTENDANCE FOR DATE: {date_query} (Total: {len(results)} students) ---")
                print(f"{'ID':<10} | {'Full Name':<25} | {'Group':<5} | {'M/A':<5} | {'Sess':<5} | {'Experiment'}")
                print("-" * 80)
                for r in results:
                    print(f"{r['student_id']:<10} | {r['student_name']:<25} | {r['group_nr']:<5} | {r['ma']:<5} | {r['session_num']:<5} | {r['experiment']}")

        elif choice == '4':
            output_csv = "danh_sach_di_lab_loc.csv"
            try:
                with open(output_csv, "w", newline="", encoding="utf-8-sig") as csvfile:
                    fieldnames = ["student_id", "student_name", "group_nr", "student_nr", "date", "day_of_week", "ma", "session_num", "experiment"]
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    
                    writer.writeheader()
                    for r in records:
                        writer.writerow(r)
                print(f"\nData successfully exported to: '{os.path.abspath(output_csv)}'")
            except Exception as e:
                print(f"\nError exporting CSV file: {e}")

        elif choice == '5':
            print("Thank you for using the program!")
            break
        else:
            print("Invalid choice, please select 1-5.")

if __name__ == "__main__":
    main()
